#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @Time     :
# @Author   : Spike713
# @Site     :
# @File     : csv_check.py
# @Purpose  :
# @Software : PyCharm
# @Copyright:
# @Licence  :

from openpyxl.reader.excel import load_workbook
# import lupa

#class Csvcheck():


def check_lua(excel_path):
        '''
        检查lua数据列,是否存在中文标点符号,是否存在不连续的{}
        column_num: lua数据列的列号码
        :return:
        '''
        wb = load_workbook(excel_path)
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])

        cn_str = ["，", "。", "；", "："]
        for i in range(2, ws.max_row + 1):
            #for j in range(2, ws.max_column + 1):
                lua_str = str(ws.cell(i, 1).value)
                for k in cn_str:
                    if k in lua_str:
                        print("检查到中文字符！！ 行数:{} 列数:{}".format(i, 1))
                if lua_str.count("{") != lua_str.count("}"):
                    print("{{和}}的个数不相等！！ 行数:{} 列数:{}".format(i, 1))

if __name__ == '__main__':
#	s = Csvcheck('/Users/Administrator/Desktop/快速生成奖励配置/reward.xlsx')
#	s.check_id()
	check_lua('/Users/Administrator/Desktop/快速生成奖励配置/reward.xlsx')