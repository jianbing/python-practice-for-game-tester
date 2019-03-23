__author__ = 'kagami'

from openpyxl.reader.excel import load_workbook

def open_workbook(excel_path:str):
	'''
	打开一个excel文件
	:param excel_path:
	:return:
	'''
	wb = load_workbook(excel_path)
	sheetnames = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheetnames[0])
	output = ""
	for i in range(1, ws.max_row):
		temp_str = "    [" + str(i) + "] = {\n"
		v_list = [j.value for j in ws[i+1] if (j.value)]
		item_type_list = v_list[::2]
		item_count_list = v_list[1::2]
		s = ""
		for i in range(len(item_type_list)):
			s = s + make_one_reward(i+1, item_type_list[i], item_count_list[i]) + '\n'
		s = temp_str + s + "    },\n"
		output = output + s
	output = "{\n" + output + "}"
	return output


def make_one_reward(row, item_type, item_count):
	s = "       [{}] = {{reward_type = REWARD_TYPE_ITEM, item_type = {}, item_count = {},}},".format(row, item_type, item_count)
	return s


s = open_workbook('/Users/kagami/git/python-practice-for-game-tester/res/快速生成奖励配置/配置模板.xlsx')
print(s)

