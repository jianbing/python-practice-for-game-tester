__author__ = 'kagami'


from openpyxl.reader.excel import load_workbook

class Csvcheck():


	def __init__(self, excel_path:str):
		'''
		打开一个excel文件
		:param excel_path:
		:return:
		'''
		wb = load_workbook(excel_path)
		sheetnames = wb.get_sheet_names()
		self.ws = wb.get_sheet_by_name(sheetnames[0])

	def check_id(self):
		'''
		检查ID是否重复 是否连续  默认ID在第一列 默认ID从1开始自增
		:return:
		'''
		id_list = []
		for i in range(2, self.ws.max_row+1):
			if self.ws.cell(i, 1).value not in id_list:
				if i - 1 != self.ws.cell(i, 1).value:
					print("ID自增错误！! 行数:{}".format(i + 1))

				id_list.append(self.ws.cell(i, 1).value)
				print(id_list)
			else:
				print("ID重复！! 行数:{}".format(i))

	def check_lua(self, column_num: int):
		'''
		检查lua数据列 是否存在中文标点符号 是否存在不连续的{}
		column_num: lua数据列的列号码
		:return:
		'''
		cn_str = ["，", "。", "；", "："]
		for i in range(2, self.ws.max_row+1):
			lua_str = str(self.ws.cell(i, column_num).value)
			for j in cn_str:
				if j in lua_str:
					print("检查到中文字符！! 行数:{} 列数:{}".format(i, column_num))
			if lua_str.count("{") != lua_str.count("}"):
				print("{{和}}的个数不相等！！ 行行数:{} 列数:{}".format(i, column_num))

if __name__ == '__main__':
	s = Csvcheck('/Users/kagami/git/python-practice-for-game-tester/res/快速生成奖励配置/配置模板.xlsx')
	s.check_id()
	s.check_lua(3)

